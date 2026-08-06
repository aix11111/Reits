"""tools.backfill_annual_nav 模块（年报年末净资产回填季度 NAV）的单元测试。

本模块复用 backfill_quarterly_nav.scan_annual_nav 解析年报缓存，核心回填逻辑
在 src.valuation.backfill_nav（另有 test_valuation 覆盖）；此处覆盖脚本自身的
两个薄封装：

- build_annual_nav：{(code, year): {nav_wan}} → {code: {year: nav_wan}}，
  None 年份剔除（早期年报未披露净值时不得进年度序列）
- write_nav_back：把重算后的 nav_wan 写回模板季度 Sheet NAV 列（覆盖式，
  NaN 写 None 清空；说明行不动）
"""

import json
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from tools.backfill_annual_nav import (
    build_annual_nav,
    update_market_quarterly_nav,
    write_nav_back,
)


def test_build_annual_nav_groups_and_drops_none():
    nav_map = {
        ("180201", 2022): {"nav_unit_price": 12.2867, "nav_wan": 860066.41},
        ("180201", 2023): {"nav_unit_price": 11.6782, "nav_wan": 817472.93},
        # 早期年报未披露净值 → None，剔除
        ("180101", 2022): {"nav_unit_price": None, "nav_wan": None},
    }

    result = build_annual_nav(nav_map)

    assert result == {
        "180201": {2022: 860066.41, 2023: 817472.93},
    }


def _build_quarterly_template(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "季度数据"
    ws.append(
        [
            "报告期",
            "基金代码",
            "基金简称",
            "营业总收入(万元)",
            "营业成本(万元)",
            "净利润(万元)",
            "可供分配金额(万元)",
            "EBITDA(万元)",
            "基金净资产-NAV(万元)",
            "数据来源/备注",
        ]
    )
    for row in rows:
        ws.append(row)
    wb.save(path)


def test_write_nav_back_overwrites_and_clears(tmp_path):
    path = tmp_path / "tpl.xlsx"
    _build_quarterly_template(
        path,
        [
            # Q4 既有期初口径值 → 覆盖为当年年报值
            ["2022Q4", "180201", "平安广州广河REIT", 1, 2, 3, 4, 5, 887511.63, "x"],
            # Q1-Q3 期初口径值仍等于上年年报值 → 覆盖同值
            ["2022Q1", "180201", "平安广州广河REIT", 1, 2, 3, 4, 5, 887511.63, "x"],
            # 无年报数据 → 清空（None）
            ["2026Q2", "508020", "某REIT", 1, 2, 3, 4, 5, 12345.0, "x"],
            # 说明行（不在回填 df）→ 保持原样
            ["说明", None, None, None, None, None, None, None, 999.0, "note"],
        ],
    )

    df = pd.DataFrame(
        [
            ["2022Q4", "180201", "平安广州广河REIT", 1, 2, 3, 4, 5, 860066.41, "x"],
            ["2022Q1", "180201", "平安广州广河REIT", 1, 2, 3, 4, 5, 887511.63, "x"],
            ["2026Q2", "508020", "某REIT", 1, 2, 3, 4, 5, float("nan"), "x"],
        ],
        columns=[
            "period",
            "code",
            "name",
            "total_revenue_wan",
            "total_cost_wan",
            "net_profit_wan",
            "distributable_wan",
            "ebitda_wan",
            "nav_wan",
            "source",
        ],
    )

    written = write_nav_back(str(path), df)

    assert written == 3
    wb = load_workbook(path)
    ws = wb["季度数据"]
    values = {ws.cell(row=r, column=9).value: ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
    assert values[860066.41] == "2022Q4"
    assert values[887511.63] == "2022Q1"
    assert None in values  # 508020 2026Q2 被清空
    # 说明行保留
    assert values.get(999.0) == "说明"


def test_update_market_quarterly_nav_uses_new_semantics(tmp_path):
    """market_quarterly.json 用新口径重算：Q4→当年年报值、Q1-Q3→上年 ffill。"""
    mq_path = tmp_path / "market_quarterly.json"
    quarters = [
        {"code": "180201", "period": "2025Q4", "nav_wan": 734851.28},
        {"code": "180201", "period": "2026Q1", "nav_wan": None},
        {"code": "508020", "period": "2026Q2", "nav_wan": None},
    ]
    mq_path.write_text(json.dumps({"quarters": quarters}), encoding="utf-8")

    annual_nav = {"180201": {2025: 700438.43, 2024: 734851.28}}

    valued = update_market_quarterly_nav(str(mq_path), annual_nav)

    assert valued == 2
    rows = json.loads(mq_path.read_text(encoding="utf-8"))["quarters"]
    by_period = {r["period"]: r for r in rows}
    # Q4 → 当年年报值（覆盖旧的去年值）
    assert by_period["2025Q4"]["nav_wan"] == pytest.approx(700438.43)
    # Q1 → 上年 ffill（最近可得年报值 = 2025 年报）
    assert by_period["2026Q1"]["nav_wan"] == pytest.approx(700438.43)
    # 无年报 → 保持 None
    assert by_period["2026Q2"]["nav_wan"] is None
