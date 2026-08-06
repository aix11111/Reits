"""季度数据 nav_wan 回填脚本（一次性数据迁移）。

季度报告不披露 NAV（基金净资产），而看板季度经营明细表展示的 NAV 列来自
年报净值。本脚本：

1. 重跑 parser_annual 于缓存年报 PDF（data/_cache/annual/ 及
   data/_cache/annual_market/），提取 (code, 报告年) → {nav_unit_price,
   nav_wan}；
2. 更新 data/market_completion.json：按 (code, year) 匹配补 nav_wan /
   nav_unit_price（None→值），缺失 (code, year) 新增行（完成度字段 None，
   name 取 market_funds）；
3. 回填模板 季度数据 Sheet 的「基金净资产-NAV(万元)」列——每季度行 NAV =
   该基金「报告年 < 季度年」的最近一份年报 nav_wan（2026Q2→2025 年报值、
   2025Q1→2024 年报值；无匹配保持空），openpyxl 仅填该列、不动结构；
4. 为 data/market_quarterly.json 每行补 nav_wan 键（同一映射逻辑）。

用法：python -m tools.reits_collector.backfill_quarterly_nav [template_xlsx]
"""

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook

from tools.reits_collector import market_fetch, parser_annual

DEFAULT_TEMPLATE = (
    Path(__file__).resolve().parents[2]
    / "data"
    / "REITsMonitor_数据模板_v1.xlsx"
)
MARKET_COMPLETION_PATH = (
    Path(__file__).resolve().parents[2] / "data" / "market_completion.json"
)
MARKET_QUARTERLY_PATH = (
    Path(__file__).resolve().parents[2] / "data" / "market_quarterly.json"
)
MARKET_FUNDS_PATH = (
    Path(__file__).resolve().parents[2] / "data" / "market_funds.json"
)
ANNUAL_CACHE_DIRS = [
    Path(__file__).resolve().parents[2] / "data" / "_cache" / "annual",
    Path(__file__).resolve().parents[2] / "data" / "_cache" / "annual_market",
]


def nav_wan_for_period(period: str, annual_nav) -> float | None:
    """季度行应回填的年报 nav_wan（万元）。

    period 形如「2026Q2」；annual_nav 为该基金年报序列，可为 {year: nav_wan}
    dict 或 [{"year", "nav_wan"}, ...] 列表（market_completion 行结构）。
    取报告年严格早于季度年的最近一份含非 None nav_wan 的年报值；
    无匹配（如 2021Q3 前无年报）返回 None。
    """
    if isinstance(annual_nav, dict):
        records = {
            int(year): nav for year, nav in annual_nav.items() if nav is not None
        }
    else:
        records = {
            int(row["year"]): row["nav_wan"]
            for row in annual_nav
            if row.get("year") is not None and row.get("nav_wan") is not None
        }
    quarter_year = int(period[:4])
    candidates = [year for year in records if year < quarter_year]
    if not candidates:
        return None
    return records[max(candidates)]


def _pdf_paths(cache_dir) -> list:
    """缓存目录全部 PDF 文件（大小写后缀兼容），按文件名排序。"""
    return sorted(
        (
            p
            for p in Path(cache_dir).iterdir()
            if p.is_file() and p.suffix.lower() == ".pdf"
        ),
        key=lambda p: p.name,
    )


def _identify_code(path, market_funds, text):
    """从文件名（沪市/模板 code 前缀）或报告标题基金全名识别基金代码。"""
    code = market_fetch._code_from_filename(path.name)
    if code is not None:
        return code
    full_name = market_fetch._report_fund_name(text)
    if full_name is None:
        return None
    szse_funds = [
        f for f in market_funds if str(f.get("code") or "").startswith("180")
    ]
    return market_fetch._match_fund_code(full_name, szse_funds)


def scan_annual_nav(cache_dirs, market_funds, errors=None):
    """扫描年报缓存目录，返回 ({(code, report_year): {nav_unit_price, nav_wan}}, errors)。

    报告年从 PDF 标题「{YYYY} 年年度报告」解析（文件名公告年 ≠ 报告年）；
    无法解析年份、无法识别代码或解析异常的文件跳过并记录 errors。
    """
    if errors is None:
        errors = []
    nav_map = {}
    for cache_dir in cache_dirs:
        for path in _pdf_paths(cache_dir):
            try:
                text = parser_annual.extract_text(path)
            except Exception as exc:
                errors.append(f"{path.name}：{exc}")
                continue
            year = parser_annual._find_report_year(text)
            if year is None:
                continue
            code = _identify_code(path, market_funds, text)
            if code is None:
                errors.append(f"{path.name}：无法识别基金代码")
                continue
            nav_map[(str(code), year)] = parser_annual._extract_nav_fields(text)
    return nav_map, errors


def _load_market_funds(path):
    """读 market_funds.json 的 funds 列表；缺失/损坏返回空列表。"""
    try:
        data = json.loads(Path(path).read_text(encoding="utf-8"))
    except (ValueError, OSError):
        return []
    return data.get("funds") if isinstance(data, dict) else []


def _load_completion(path):
    """读 market_completion.json 的 completion 行；缺失/损坏返回空列表。"""
    try:
        data = json.loads(Path(path).read_text(encoding="utf-8"))
    except (ValueError, OSError):
        return []
    completion = data.get("completion") if isinstance(data, dict) else None
    return completion if isinstance(completion, list) else []


def update_market_completion(mc_path, nav_map, market_funds):
    """按 (code, year) 更新 market_completion.json 净值字段并写回。

    已存在行补 None→值（非 None 不覆盖）；缺失 (code, year) 新增行
    （完成度字段 None，name 取 market_funds；code 不在 funds 则跳过）。
    返回 (filled, added)：filled 为补值行数，added 为新增行数。
    """
    path = Path(mc_path)
    completion = _load_completion(path)
    names = {str(f.get("code")): str(f.get("name") or f.get("code")) for f in market_funds}
    by_key = {(str(r.get("code")), int(r["year"])): r for r in completion if r.get("year") is not None}

    filled = 0
    added = 0
    for (code, year), nav in nav_map.items():
        row = by_key.get((code, year))
        if row is not None:
            for key in ("nav_unit_price", "nav_wan"):
                if row.get(key) is None and nav.get(key) is not None:
                    row[key] = nav[key]
                    filled += 1
            continue
        if code not in names:
            continue
        new_row = {
            "code": code,
            "name": names[code],
            "year": year,
            "predicted_wan": None,
            "actual_wan": None,
            "completion_pct": None,
            "nav_unit_price": nav.get("nav_unit_price"),
            "nav_wan": nav.get("nav_wan"),
        }
        completion.append(new_row)
        by_key[(code, year)] = new_row
        added += 1

    completion.sort(key=lambda r: (int(r.get("year") or -1), str(r.get("code") or "")))
    path.write_text(
        json.dumps({"completion": completion}, ensure_ascii=False, indent=1),
        encoding="utf-8",
    )
    return filled, added


def annual_nav_by_code(completion) -> dict:
    """completion 行 → {code: {year: nav_wan}}（仅保留非 None nav_wan）。"""
    by_code = {}
    for row in completion:
        if row.get("nav_wan") is None or row.get("year") is None:
            continue
        by_code.setdefault(str(row.get("code")), {})[int(row["year"])] = row["nav_wan"]
    return by_code


def update_template_nav(template_path, annual_nav_by_code):
    """按 (code, period) 把 nav_wan（万元）写入模板季度 Sheet 的 NAV 列。

    NAV 列固定按表头含「NAV」定位（表头含换行，避免字符串匹配漂移）。
    无匹配行保持原样。返回实际回填行数。
    """
    wb = load_workbook(template_path)
    ws = wb["季度数据"]
    nav_col = None
    for idx, cell in enumerate(ws[1], start=1):
        if "NAV" in str(cell.value or ""):
            nav_col = idx
            break
    if nav_col is None:
        raise ValueError("模板季度数据 Sheet 未找到「NAV」列")

    filled = 0
    for row in ws.iter_rows(min_row=2):
        code = row[1].value
        period = row[0].value
        if code is None or period is None:
            continue
        nav = nav_wan_for_period(
            str(period), annual_nav_by_code.get(str(code), {})
        )
        if nav is not None:
            row[nav_col - 1].value = nav
            filled += 1
    wb.save(template_path)
    return filled


def update_market_quarterly_nav(mq_path, annual_nav_by_code):
    """为 market_quarterly.json 每行补 nav_wan 键并写回。

    无匹配行写 None。返回带值（非 None）的行数。
    """
    path = Path(mq_path)
    data = json.loads(path.read_text(encoding="utf-8"))
    quarters = data.get("quarters") or []
    valued = 0
    for row in quarters:
        nav = nav_wan_for_period(
            str(row.get("period") or ""),
            annual_nav_by_code.get(str(row.get("code")), {}),
        )
        row["nav_wan"] = nav
        if nav is not None:
            valued += 1
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8"
    )
    return valued


def main(argv=None):
    """CLI：扫描年报缓存 → 更新 market_completion.json → 回填模板与
    market_quarterly.json，打印摘要。"""
    parser = argparse.ArgumentParser(description="回填季度 nav_wan")
    parser.add_argument(
        "template",
        nargs="?",
        default=str(DEFAULT_TEMPLATE),
        help="模板 xlsx 路径（默认 data/REITsMonitor_数据模板_v1.xlsx）",
    )
    args = parser.parse_args(argv)

    market_funds = _load_market_funds(MARKET_FUNDS_PATH)
    nav_map, errors = scan_annual_nav(ANNUAL_CACHE_DIRS, market_funds)
    filled, added = update_market_completion(MARKET_COMPLETION_PATH, nav_map, market_funds)

    completion = _load_completion(MARKET_COMPLETION_PATH)
    nav_by_code = annual_nav_by_code(completion)
    template_filled = update_template_nav(args.template, nav_by_code)
    mq_valued = update_market_quarterly_nav(MARKET_QUARTERLY_PATH, nav_by_code)

    print(f"年报缓存识别净值 {len(nav_map)} 条")
    print(f"market_completion.json 补值 {filled} 行、新增 {added} 行")
    print(f"模板季度 Sheet 回填 {template_filled} 行")
    print(f"market_quarterly.json 有值行 {mq_valued} 行")
    for err in errors:
        print(f"  [错误] {err}")
    return {
        "scanned": len(nav_map),
        "filled": filled,
        "added": added,
        "template_filled": template_filled,
        "mq_valued": mq_valued,
    }


if __name__ == "__main__":
    main()
