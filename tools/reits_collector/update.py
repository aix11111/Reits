"""增量数据更新管线。

流程：读取模板月度/季度数据 → 逐基金（14 只 REIT）按市场拉公告
（深市 180xxx 走 cninfo，沪市 508xxx 走 sse）→ 按标题过滤
（月度含「运营数据」，季度含「季度报告」且不含「提示性」）→ 下载
PDF 到临时目录（已存在跳过）→ 解析 → 跳过已存在 (code, period) →
新行追加回模板对应 Sheet（period 升序 + code）。

年报可供分配完成度（fetch_new_completion）：按标题过滤（含「年度报告」
且不含「摘要」「提示性」）→ 解析 parser_annual.parse_annual_completion →
跳过已存在 (code, year) → 新记录合并写回 annual_completion.json。
沪市列表失败 time.sleep(90) 重试 3 次（限流防护），沪市基金间 15s 间隔。

单只基金失败只记录到 errors，不影响其余基金。
"""

import argparse
import json
import numbers
import shutil
import tempfile
import time
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from src import data_loader, market_data
from tools.reits_collector import (
    cninfo,
    parser_annual,
    parser_generic,
    parser_quarterly,
    sse,
)

FUND_CODES = [
    "180201",
    "180202",
    "180203",
    "508001",
    "508018",
    "508008",
    "508066",
    "508009",
    "508007",
    "508033",
    "508086",
    "508069",
    "508036",
    "508020",
]

MONTHLY_SOURCE = "月度运营公告（自动采集）"
QUARTERLY_SOURCE = "季度报告（自动采集）"
CNINFO_PDF_BASE = "https://static.cninfo.com.cn/"
DEFAULT_TEMPLATE = (
    Path(__file__).resolve().parents[2]
    / "data"
    / "REITsMonitor_数据模板_v1.xlsx"
)

# PDF 下载临时目录：由 main() 创建并在结束后清理；测试中由测试替身替换。
PDF_DIR = None

# 市值快照与份额数据文件路径
SNAPSHOT_PATH = Path(__file__).resolve().parents[2] / "data" / "market_snapshot.json"
FUND_SHARES_PATH = Path(__file__).resolve().parents[2] / "data" / "fund_shares.json"


def _period_date_from(period: str, kind: str) -> str:
    """最新报告期 → 查询起始日。

    月度 YYYY-MM → 次月 1 日；季度 YYYYQq → 下一季度首日。
    """
    if kind == "monthly":
        year, month = period.split("-")
        month = int(month)
        if month == 12:
            return f"{int(year) + 1}-01-01"
        return f"{year}-{month + 1:02d}-01"
    year, quarter = period[:4], int(period[5])
    month = quarter * 3 + 1
    if month > 12:
        return f"{int(year) + 1}-01-01"
    return f"{year}-{month:02d}-01"


def _date_from(code: str, last_period: dict, kind: str) -> str:
    """逐基金查询起始日：无数据从 2023-01-01 起；否则从该基金最新报告期之后起。"""
    last = last_period.get(code)
    if last is None:
        return "2023-01-01"
    return _period_date_from(last, kind)


def _list_for(code: str, date_from: str, date_to: str) -> list:
    """深市走 cninfo（先检索 orgId），沪市走 sse。"""
    if code.startswith("180"):
        org_id = cninfo.search_org_id(code)
        return cninfo.list_announcements(code, org_id, date_from, date_to)
    return sse.list_announcements(code, date_from, date_to)


def _download(url_path: str, dest: Path, code: str) -> None:
    """按市场拼接完整 URL 并下载；cninfo 用静态域名前缀，sse 由接口补全。"""
    if code.startswith("180"):
        cninfo.download_pdf(CNINFO_PDF_BASE + url_path, dest)
    else:
        sse.download_pdf(url_path, dest)


def _announcement_rows(df, kind, errors, title_filter, parse_fn, source):
    """通用抓取管线：逐基金拉公告→过滤→下载→解析→去重→组装新行。"""
    if errors is None:
        errors = []
    existing = set()
    code_to_name = {}
    last_period = {}
    for _, record in df.iterrows():
        code = str(record["code"])
        period = str(record["period"])
        existing.add((code, period))
        code_to_name[code] = str(record["name"])
        if period > last_period.get(code, ""):
            last_period[code] = period
    date_to = date.today().isoformat()

    rows = []
    for code in FUND_CODES:
        try:
            announcements = _list_for(code, _date_from(code, last_period, kind), date_to)
        except Exception as exc:
            errors.append(f"{code}：{exc}")
            continue
        for item in announcements:
            try:
                title = item.get("announcementTitle") or item.get("title") or ""
                if not title_filter(title):
                    continue
                url_path = item.get("adjunctUrl") or item.get("url") or ""
                if not url_path:
                    continue
                dest = PDF_DIR / Path(url_path).name
                if not dest.exists():
                    _download(url_path, dest, code)
                parsed = parse_fn(dest)
                period = parsed.get("period")
                if period is None or (code, period) in existing:
                    continue
                rows.append(
                    dict(
                        parsed,
                        code=code,
                        name=code_to_name.get(code, code),
                        source=source,
                    )
                )
                existing.add((code, period))
            except Exception as exc:
                errors.append(f"{code}：{exc}")
    return rows


def fetch_new_monthly(monthly_df, errors=None):
    """抓取月度运营公告，返回新行 dict 列表（含 code/name/period/source）。"""
    return _announcement_rows(
        monthly_df,
        "monthly",
        errors,
        title_filter=lambda title: "运营数据" in title,
        parse_fn=parser_generic.parse_pdf,
        source=MONTHLY_SOURCE,
    )


def fetch_new_quarterly(quarterly_df, errors=None):
    """抓取季度报告，返回新行 dict 列表（含 code/name/period/source）。"""
    return _announcement_rows(
        quarterly_df,
        "quarterly",
        errors,
        title_filter=lambda title: "季度报告" in title and "提示性" not in title,
        parse_fn=parser_quarterly.parse_quarterly_report,
        source=QUARTERLY_SOURCE,
    )


def _annual_title_filter(title):
    """年报标题过滤：含「年度报告」，排除「摘要」与「提示性」。"""
    return "年度报告" in title and "摘要" not in title and "提示性" not in title


def _completion_date_from(existing, code) -> str:
    """年报完成度逐基金查询起始日：无记录从 2023-01-01 起；
    否则从该基金最新 year 次年 1 月 1 日起（当年年报尚未披露）。"""
    years = [int(y) for c, y in existing if c == code]
    if not years:
        return "2023-01-01"
    return f"{max(years) + 1}-01-01"


def _sse_list_with_retry(code, date_from, date_to, attempts=3, delay=90):
    """沪市列表失败 time.sleep(delay) 重试 attempts 次（限流防护）；
    仍失败抛最后一次异常。"""
    last_exc = None
    for attempt in range(attempts):
        try:
            return sse.list_announcements(code, date_from, date_to)
        except Exception as exc:
            last_exc = exc
            if attempt < attempts - 1:
                time.sleep(delay)
    raise last_exc


def fetch_new_completion(lst, code_to_name=None, errors=None):
    """抓取年报可供分配完成度，返回新记录 dict 列表。

    lst：已存在的完成度记录（含 code/year/name，如 annual_completion.json 的
    completion 数组）。已存在 (code, year) 跳过；新 year 记录（year/
    predicted_wan/actual_wan/completion_pct + code/name）追加。
    沪市列表失败 90s 重试 3 次；沪市基金间 15s 间隔；下载复用（已存在跳过）。
    """
    if errors is None:
        errors = []
    existing = set()
    names = dict(code_to_name) if code_to_name else {}
    for record in lst:
        code = str(record.get("code")) if record.get("code") is not None else ""
        year = record.get("year")
        if code and year is not None:
            existing.add((code, str(int(year))))
        name = record.get("name")
        if code and name and code not in names:
            names[code] = str(name)
    date_to = date.today().isoformat()

    rows = []
    for idx, code in enumerate(FUND_CODES):
        is_sh = not code.startswith("180")
        try:
            date_from = _completion_date_from(existing, code)
            if is_sh:
                announcements = _sse_list_with_retry(code, date_from, date_to)
            else:
                org_id = cninfo.search_org_id(code)
                announcements = cninfo.list_announcements(
                    code, org_id, date_from, date_to
                )
        except Exception as exc:
            errors.append(f"{code}：{exc}")
            announcements = []
        if is_sh and idx != len(FUND_CODES) - 1:
            time.sleep(15)
        for item in announcements:
            try:
                title = item.get("announcementTitle") or item.get("title") or ""
                if not _annual_title_filter(title):
                    continue
                url_path = item.get("adjunctUrl") or item.get("url") or ""
                if not url_path:
                    continue
                dest = PDF_DIR / Path(url_path).name
                if not dest.exists():
                    _download(url_path, dest, code)
                parsed = parser_annual.parse_annual_completion(dest)
                year = parsed.get("year")
                if year is None or (code, str(int(year))) in existing:
                    continue
                rows.append(dict(parsed, code=code, name=names.get(code, code)))
                existing.add((code, str(int(year))))
            except Exception as exc:
                errors.append(f"{code}：{exc}")
    return rows


def _code_to_name_from_df(monthly_df):
    """从月度 DataFrame 提取 code → name 映射（完成度新记录兜底命名用）。"""
    code_to_name = {}
    for _, record in monthly_df.iterrows():
        code = str(record["code"])
        name = str(record["name"])
        if code not in code_to_name:
            code_to_name[code] = name
    return code_to_name


def _to_native(value):
    """DataFrame 取值 → openpyxl 可写类型；NaN → None。"""
    if value is None:
        return None
    if isinstance(value, float):
        if value != value:
            return None
        return value
    if isinstance(value, numbers.Integral):
        return int(value)
    if isinstance(value, numbers.Real):
        return float(value)
    return value


def _monthly_row_values(row):
    return [
        row.get("period"),
        row.get("code"),
        row.get("name"),
        row.get("toll_revenue_wan"),
        row.get("daily_traffic"),
        row.get("toll_revenue_yoy", row.get("revenue_yoy")),
        row.get("traffic_yoy"),
        row.get("source") or MONTHLY_SOURCE,
    ]


def _quarterly_row_values(row):
    return [
        row.get("period"),
        row.get("code"),
        row.get("name"),
        row.get("total_revenue_wan", row.get("revenue_wan")),
        row.get("total_cost_wan"),
        row.get("net_profit_wan"),
        row.get("distributable_wan"),
        row.get("ebitda_wan"),
        row.get("nav_wan"),
        row.get("source") or QUARTERLY_SOURCE,
    ]


def _rewrite_sheet(wb, sheet_name, old_df, new_rows, row_values_fn):
    """删掉数据行（行 2~max_row），重写 旧+新 合并行，period 升序 + code。"""
    ws = wb[sheet_name]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    all_rows = []
    for _, record in old_df.iterrows():
        all_rows.append(
            row_values_fn({col: _to_native(record[col]) for col in old_df.columns})
        )
    for record in new_rows:
        all_rows.append(row_values_fn({k: _to_native(v) for k, v in record.items()}))
    all_rows.sort(key=lambda values: (str(values[0]), str(values[1])))
    for values in all_rows:
        ws.append(values)


def _update_completion_file(completion_path, code_to_name, errors):
    """读 annual_completion.json → fetch_new_completion 抓取 → 合并写回，
    返回新增记录列表。文件缺失/损坏按空处理；无新增不写回。"""
    completion_path = Path(completion_path)
    try:
        data = json.loads(completion_path.read_text(encoding="utf-8"))
    except (ValueError, OSError):
        data = {}
    existing = data.get("completion") or []
    new_rows = fetch_new_completion(existing, code_to_name=code_to_name, errors=errors)
    if new_rows:
        data["completion"] = list(existing) + new_rows
        completion_path.write_text(
            json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8"
        )
    return new_rows


def _load_snapshot():
    """读 SNAPSHOT_PATH；文件缺失/损坏返回空结构 {"snapshots": [], "latest": {}}。"""
    try:
        data = json.loads(SNAPSHOT_PATH.read_text(encoding="utf-8"))
    except (ValueError, OSError):
        data = {}
    if not isinstance(data, dict):
        data = {}
    return {
        "snapshots": data.get("snapshots") or [],
        "latest": data.get("latest") or {},
    }


def _load_fund_shares():
    """读 data/fund_shares.json 的 shares 映射（code → 份额）；缺失/损坏返回空 dict。"""
    try:
        data = json.loads(FUND_SHARES_PATH.read_text(encoding="utf-8"))
    except (ValueError, OSError):
        data = {}
    shares = data.get("shares") if isinstance(data, dict) else None
    return shares if isinstance(shares, dict) else {}


def fetch_market_snapshot(shares, errors=None):
    """抓取全市场 REIT 实时价生成市值快照，追加历史并按 date+code 去重。

    shares：code → 份额总额 映射。成功基金新增当日快照行（price × shares /
    10000，round 2）并覆盖 latest；失败基金沿用旧快照 latest 并记录 errors。
    全失败（行情为空）返回旧快照内容不变；无旧快照时返回空结构，均不抛错。
    """
    if errors is None:
        errors = []
    old = _load_snapshot()
    old_snapshots = old["snapshots"]
    old_latest = old["latest"]
    today = date.today().isoformat()

    quotes = market_data.get_realtime_quotes()
    prices = {}
    if quotes is not None:
        for _, row in quotes.iterrows():
            try:
                price = float(row["price"])
            except (TypeError, ValueError):
                continue
            if price == price:
                prices[str(row["code"])] = price

    new_snapshots = []
    latest = dict(old_latest)
    for code in FUND_CODES:
        if code not in prices:
            errors.append(f"{code}：实时行情缺失")
            continue
        if shares.get(code) is None:
            errors.append(f"{code}：份额缺失")
            continue
        price = prices[code]
        market_cap_wan = round(price * shares[code] / 10000, 2)
        new_snapshots.append(
            {
                "date": today,
                "code": code,
                "price": price,
                "market_cap_wan": market_cap_wan,
            }
        )
        latest[code] = {"price": price, "market_cap_wan": market_cap_wan}

    if not new_snapshots:
        return {"snapshots": old_snapshots, "latest": old_latest}

    seen = {(row["date"], row["code"]) for row in old_snapshots}
    snapshots = old_snapshots + [
        row for row in new_snapshots if (row["date"], row["code"]) not in seen
    ]
    snapshots.sort(key=lambda row: (row["date"], row["code"]))
    result = {"snapshots": snapshots, "latest": latest}
    SNAPSHOT_PATH.write_text(
        json.dumps(result, ensure_ascii=False, indent=1), encoding="utf-8"
    )
    return result


def update_template(template_path, completion_path=None):
    """读取模板 → 抓取增量 → 重写月度/季度 Sheet；completion_path 非 None 时
    同步更新 annual_completion.json；市值快照追加最新收盘价。
    返回摘要 dict（含 completion_added / snapshot_updated）。"""
    errors = []
    monthly_df = data_loader.load_monthly(template_path)
    quarterly_df = data_loader.load_quarterly(template_path)
    monthly_new = fetch_new_monthly(monthly_df, errors=errors)
    quarterly_new = fetch_new_quarterly(quarterly_df, errors=errors)

    completion_new = []
    if completion_path is not None:
        completion_new = _update_completion_file(
            completion_path, _code_to_name_from_df(monthly_df), errors
        )

    snapshot_result = fetch_market_snapshot(_load_fund_shares(), errors=errors)
    today = date.today().isoformat()
    snapshot_updated = any(
        row.get("date") == today for row in snapshot_result.get("snapshots", [])
    )

    wb = load_workbook(template_path)
    _rewrite_sheet(wb, "月度数据", monthly_df, monthly_new, _monthly_row_values)
    _rewrite_sheet(wb, "季度数据", quarterly_df, quarterly_new, _quarterly_row_values)
    wb.save(template_path)

    return {
        "monthly_added": len(monthly_new),
        "quarterly_added": len(quarterly_new),
        "completion_added": len(completion_new),
        "snapshot_updated": snapshot_updated,
        "errors": errors,
    }


def main(argv=None):
    """CLI：增量更新模板并打印摘要；PDF 临时目录用完清理。"""
    parser = argparse.ArgumentParser(description="增量更新 REITs 模板 Excel")
    parser.add_argument(
        "template",
        nargs="?",
        default=str(DEFAULT_TEMPLATE),
        help="模板 xlsx 路径（默认 data/REITsMonitor_数据模板_v1.xlsx）",
    )
    parser.add_argument(
        "--completion",
        default=str(DEFAULT_TEMPLATE.parent / "annual_completion.json"),
        help="annual_completion.json 路径",
    )
    args = parser.parse_args(argv)

    global PDF_DIR
    PDF_DIR = Path(tempfile.mkdtemp(prefix="reits_pdf_"))
    try:
        summary = update_template(args.template, completion_path=args.completion)
    finally:
        shutil.rmtree(PDF_DIR, ignore_errors=True)

    print(
        f"月度新增 {summary['monthly_added']} 行，"
        f"季度新增 {summary['quarterly_added']} 行，"
        f"年度完成度新增 {summary['completion_added']} 条"
    )
    for err in summary["errors"]:
        print(f"  [错误] {err}")
    return summary


if __name__ == "__main__":
    main()
