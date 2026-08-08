import re

import pandas as pd
import pytest
from openpyxl import Workbook

from src.data_loader import (
    STATIC_COLS,
    MONTHLY_COLS,
    QUARTERLY_COLS,
    load_static,
    load_monthly,
    load_quarterly,
    load_all,
)

STATIC_HEADERS = [
    "基金代码",
    "基金简称",
    "底层资产",
    "区域",
    "里程(km)",
    "上市日期",
    "发行规模(亿元)",
    "特许经营剩余年限\n(截至2026)",
    "资产类型",
]

STATIC_ROWS = [
    ["180201", "平安广州广河REIT", "广河高速(广州段)", "华南", 99, "2021-06-07", 91.14, 9, "高速公路"],
    ["180202", "华夏越秀高速REIT", "汉孝高速", "华中", 34, "2021-12-03", 21.3, 11.5, "高速公路"],
]

MONTHLY_HEADERS = [
    "报告期",
    "基金代码",
    "基金简称",
    "通行费收入(万元)",
    "日均自然车流量(辆/日)",
    "通行费收入同比(%)",
    "车流量同比(%)",
    "数据来源/备注",
]

MONTHLY_ROWS = [
    ["2025-01", "180201", "平安广州广河REIT", 8256, 128500, 2.3, 1.8, "月度运营公告"],
    ["2025-02", "180201", "平安广州广河REIT", 7200, 115000, -1.5, -2.1, "月度运营公告"],
]

QUARTERLY_HEADERS = [
    "报告期",
    "基金代码",
    "基金简称",
    "营业总收入(万元)",
    "营业成本(万元)",
    "净利润(万元)",
    "可供分配金额(万元)",
    "EBITDA(万元)",
    "基金净资产-NAV(万元)",
    "数据来源/备注",
]

QUARTERLY_ROWS = [
    ["2025Q1", "180201", "平安广州广河REIT", 24500, 8500, 6800, 5200, 15800, 685000, "2025年一季报"],
    ["2025Q2", "180201", "平安广州广河REIT", 25000, 8700, 7000, 5400, 16200, 688000, "2025年中报"],
]


@pytest.fixture
def xlsx_path(tmp_path):
    path = tmp_path / "test_data.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("静态信息")
    ws.append(STATIC_HEADERS)
    for row in STATIC_ROWS:
        ws.append(row)
    ws.append(["📝 填写说明：", None, None, None, None, None, None, None, None])

    ws = wb.create_sheet("月度数据")
    ws.append(MONTHLY_HEADERS)
    for row in MONTHLY_ROWS:
        ws.append(row)
    ws.append(["填写说明", None, None, None, None, None, None, None])

    ws = wb.create_sheet("季度数据")
    ws.append(QUARTERLY_HEADERS)
    for row in QUARTERLY_ROWS:
        ws.append(row)
    ws.append(["填写说明", None, None, None, None, None, None, None, None, None])

    wb.save(path)
    return path


def test_constant_column_counts():
    assert len(STATIC_COLS) == 9
    assert len(MONTHLY_COLS) == 8
    assert len(QUARTERLY_COLS) == 10


def test_constant_static_maps_concession_header_with_newline():
    assert STATIC_COLS["特许经营剩余年限\n(截至2026)"] == "concession_years_left"
    assert STATIC_COLS["里程(km)"] == "mileage_km"
    assert STATIC_COLS["发行规模(亿元)"] == "issue_scale_yi"


def test_load_static_normalizes_columns_and_values(xlsx_path):
    df = load_static(xlsx_path)

    assert list(df.columns) == [
        "code",
        "name",
        "asset",
        "region",
        "mileage_km",
        "listing_date",
        "issue_scale_yi",
        "concession_years_left",
        "asset_type",
    ]
    assert df["code"].tolist() == ["180201", "180202"]
    assert df["name"].tolist() == ["平安广州广河REIT", "华夏越秀高速REIT"]
    assert df["asset"].tolist() == ["广河高速(广州段)", "汉孝高速"]
    assert df["region"].tolist() == ["华南", "华中"]
    assert df["mileage_km"].tolist() == [99, 34]
    assert df["listing_date"].tolist() == ["2021-06-07", "2021-12-03"]
    assert df["issue_scale_yi"].tolist() == pytest.approx([91.14, 21.3])
    assert df["concession_years_left"].tolist() == pytest.approx([9, 11.5])
    assert df["asset_type"].tolist() == ["高速公路", "高速公路"]


def test_load_monthly_normalizes_columns_and_values(xlsx_path):
    df = load_monthly(xlsx_path)

    assert list(df.columns) == [
        "period",
        "code",
        "name",
        "toll_revenue_wan",
        "daily_traffic",
        "toll_revenue_yoy",
        "traffic_yoy",
        "source",
    ]
    assert df["period"].tolist() == ["2025-01", "2025-02"]
    assert df["code"].tolist() == ["180201", "180201"]
    assert df["toll_revenue_wan"].tolist() == [8256, 7200]
    assert df["daily_traffic"].tolist() == [128500, 115000]
    assert df["toll_revenue_yoy"].tolist() == pytest.approx([2.3, -1.5])
    assert df["traffic_yoy"].tolist() == pytest.approx([1.8, -2.1])
    assert df["source"].tolist() == ["月度运营公告", "月度运营公告"]


def test_load_quarterly_normalizes_columns_and_values(xlsx_path):
    df = load_quarterly(xlsx_path)

    assert list(df.columns) == [
        "period",
        "code",
        "name",
        "total_revenue_wan",
        "total_cost_wan",
        "net_profit_wan",
        "distributable_wan",
        "ebitda_wan",
        "nav_wan",
        "source",
    ]
    assert df["period"].tolist() == ["2025Q1", "2025Q2"]
    assert df["total_revenue_wan"].tolist() == [24500, 25000]
    assert df["total_cost_wan"].tolist() == [8500, 8700]
    assert df["net_profit_wan"].tolist() == [6800, 7000]
    assert df["distributable_wan"].tolist() == [5200, 5400]
    assert df["ebitda_wan"].tolist() == [15800, 16200]
    assert df["nav_wan"].tolist() == [685000, 688000]
    assert df["source"].tolist() == ["2025年一季报", "2025年中报"]


def test_load_all_returns_three_dataframes(xlsx_path):
    result = load_all(xlsx_path)

    assert set(result.keys()) == {"static", "monthly", "quarterly"}
    for key in ("static", "monthly", "quarterly"):
        assert isinstance(result[key], pd.DataFrame)
        assert not result[key].empty
    assert list(result["static"].columns)[0] == "code"
    assert list(result["monthly"].columns)[0] == "period"
    assert list(result["quarterly"].columns)[0] == "period"


@pytest.mark.parametrize(
    "loader", [load_static, load_monthly, load_quarterly, load_all]
)
def test_load_filters_out_instruction_rows(xlsx_path, loader):
    result = loader(xlsx_path)
    if isinstance(result, dict):
        for df in result.values():
            assert len(df) == 2
    else:
        assert len(result) == 2


@pytest.mark.parametrize(
    "loader", [load_static, load_monthly, load_quarterly, load_all]
)
def test_missing_file_raises_file_not_found_with_path(tmp_path, loader):
    missing = tmp_path / "missing.xlsx"
    with pytest.raises(FileNotFoundError) as excinfo:
        loader(missing)
    assert str(missing) in str(excinfo.value)


# ---------------------------------------------------------------------------
# 美国数据加载（load_us_funds / load_us_annual / load_us_snapshot，复用 load_sg 模式）
# ---------------------------------------------------------------------------


def test_load_us_funds_maps_ticker_to_name(tmp_path):
    import json

    from src.data_loader import load_us_funds

    path = tmp_path / "us_funds.json"
    path.write_text(
        json.dumps(
            {
                "funds": [
                    {"ticker": "PLD", "name": "Prologis"},
                    {"ticker": "O", "name": "Realty Income"},
                ]
            }
        ),
        encoding="utf-8",
    )

    result = load_us_funds(path)

    assert result == {"PLD": "Prologis", "O": "Realty Income"}


def test_load_us_annual_groups_by_ticker(tmp_path):
    import json

    from src.data_loader import load_us_annual

    path = tmp_path / "us_annual.json"
    path.write_text(
        json.dumps(
            {
                "annual": [
                    {"ticker": "PLD", "fiscal_year": "2025", "dpu_usd": 4.04},
                    {"ticker": "PLD", "fiscal_year": "2024", "dpu_usd": 3.75},
                    {"ticker": "O", "fiscal_year": "2025", "dpu_usd": 3.24},
                ]
            }
        ),
        encoding="utf-8",
    )

    result = load_us_annual(path)

    assert set(result.keys()) == {"PLD", "O"}
    assert len(result["PLD"]) == 2
    assert result["PLD"][0]["dpu_usd"] == 4.04


def test_load_us_snapshot_returns_latest(tmp_path):
    import json

    from src.data_loader import load_us_snapshot

    path = tmp_path / "us_market_snapshot.json"
    path.write_text(
        json.dumps({"latest": {"PLD": 140.16, "O": 62.51}}),
        encoding="utf-8",
    )

    result = load_us_snapshot(path)

    assert result == {"PLD": 140.16, "O": 62.51}


def test_load_us_missing_files_return_empty(tmp_path):
    from src.data_loader import load_us_annual, load_us_funds, load_us_snapshot

    missing = tmp_path / "missing.json"
    assert load_us_funds(missing) == {}
    assert load_us_annual(missing) == {}
    assert load_us_snapshot(missing) == {}
