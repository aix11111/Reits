"""tools.reits_collector.update 模块（增量数据更新管线）的单元测试。

通过 monkeypatch 替换 cninfo / sse / parser_generic / parser_quarterly /
parser_annual 的列表、下载与解析函数，全程不发起真实网络请求。覆盖：
- 已存在 period 跳过、新 period 追加（月度 + 季度 + 年报完成度）
- 深市走 cninfo、沪市走 sse 的分支路由
- 单基金网络失败 → 收集到 errors 不崩溃
- 沪市列表失败 time.sleep(90) 重试 3 次（限流防护）、沪市基金间 15s 间隔
- 年报「摘要」「提示性」公告过滤、下载复用（已存在 PDF 跳过）
- update_template 写回模板后 load 验证行数增加、写回 annual_completion.json
- 空增量（无新公告）→ 摘要全 0、completion_added=0
"""

import json
from datetime import date
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from src import data_loader
from tools.reits_collector import (
    cninfo,
    parser_annual,
    parser_generic,
    parser_quarterly,
    sse,
    update,
)

MONTHLY_NUMERIC_KEYS = [
    "daily_traffic",
    "traffic_mom",
    "traffic_yoy",
    "traffic_cum",
    "traffic_cum_yoy",
    "toll_revenue_wan",
    "revenue_mom",
    "revenue_yoy",
    "revenue_cum",
    "revenue_cum_yoy",
]

SH_CODES = [
    "508001",
    "508018",
    "508008",
    "508066",
    "508009",
    "508007",
    "508033",
    "508086",
    "508069",
    "508036",
    "508020",
]
SZ_CODES = ["180201", "180202", "180203"]


def _monthly_parse_result(period):
    """构造 parse_pdf 风格的解析结果：period + project_name + 10 个数值字段。"""
    result = {"period": period, "project_name": "某高速"}
    for i, key in enumerate(MONTHLY_NUMERIC_KEYS):
        result[key] = 1000 + i
    return result


def _monthly_df(rows):
    return pd.DataFrame(
        rows,
        columns=[
            "period",
            "code",
            "name",
            "toll_revenue_wan",
            "daily_traffic",
            "toll_revenue_yoy",
            "traffic_yoy",
            "source",
        ],
    )


def _quarterly_df(rows):
    return pd.DataFrame(
        rows,
        columns=[
            "period",
            "code",
            "name",
            "total_revenue_wan",
            "total_cost_wan",
            "net_profit_wan",
            "distributable_wan",
            "ebitda_wan",
            "nav_wan",
            "source",
        ],
    )


def _patch_downloads(monkeypatch, tmp_path):
    """PDF 目录指向 tmp_path；两个市场的 download_pdf 均为写假 PDF 字节。"""
    monkeypatch.setattr(update, "PDF_DIR", tmp_path)

    def fake_download(url, dest):
        dest.write_bytes(b"%PDF-fake")
        return dest

    monkeypatch.setattr(cninfo, "download_pdf", fake_download)
    monkeypatch.setattr(sse, "download_pdf", fake_download)


def _cninfo_item(title, filename):
    return {
        "announcementTitle": title,
        "adjunctUrl": f"finalpage/2026-07-15/{filename}",
        "announcementTime": 1752566400000,
    }


def _sse_item(title, filename):
    return {
        "title": title,
        "url": f"/disclosure/fund/announcement/{filename}",
        "sseDate": "2026-07-15",
    }


# ---------------------------------------------------------------------------
# fetch_new_monthly
# ---------------------------------------------------------------------------


def test_fetch_new_monthly_skips_existing_period_and_appends_new(
    monkeypatch, tmp_path
):
    """已存在 (code, period) 跳过、新 period 追加；非「运营数据」公告被过滤；
    已存在的 PDF 跳过下载但仍解析。"""
    _patch_downloads(monkeypatch, tmp_path)
    monthly_df = _monthly_df(
        [["2026-06", "180201", "平安广州广河REIT", 8000, 120000, 1.0, 2.0, "公告"]]
    )
    # 2026-07 的 PDF 已存在 → 应跳过下载
    (tmp_path / "180201_202607.pdf").write_bytes(b"%PDF-exists")

    monkeypatch.setattr(cninfo, "search_org_id", lambda code: f"org-{code}")
    monkeypatch.setattr(sse, "list_announcements", lambda *a, **k: [])

    def fake_cninfo_list(code, org_id, date_from, date_to, page_size=100):
        if code != "180201":
            return []
        return [
            _cninfo_item("关于2026年6月主要运营数据的公告", "180201_202606.pdf"),
            _cninfo_item("关于2026年7月主要运营数据的公告", "180201_202607.pdf"),
            _cninfo_item("关于召开基金份额持有人大会的公告", "other.pdf"),
        ]

    monkeypatch.setattr(cninfo, "list_announcements", fake_cninfo_list)

    downloaded = []
    real_download = cninfo.download_pdf

    def recording_download(url, dest):
        downloaded.append(dest.name)
        return real_download(url, dest)

    monkeypatch.setattr(cninfo, "download_pdf", recording_download)

    parsed = {
        "180201_202606.pdf": _monthly_parse_result("2026-06"),
        "180201_202607.pdf": _monthly_parse_result("2026-07"),
    }
    parse_calls = []

    def fake_parse(path):
        parse_calls.append(path.name)
        return parsed[path.name]

    monkeypatch.setattr(parser_generic, "parse_pdf", fake_parse)

    rows = update.fetch_new_monthly(monthly_df)

    assert len(rows) == 1
    row = rows[0]
    assert row["code"] == "180201"
    assert row["name"] == "平安广州广河REIT"
    assert row["period"] == "2026-07"
    for key in MONTHLY_NUMERIC_KEYS:
        assert key in row

    # 已存在 period 的公告仍下载解析（用于判定 period）；已存在 PDF 跳过下载；
    # 非「运营数据」公告既不下载也不解析
    assert downloaded == ["180201_202606.pdf"]
    assert sorted(parse_calls) == ["180201_202606.pdf", "180201_202607.pdf"]


def test_fetch_new_monthly_routes_sz_to_cninfo_and_sh_to_sse(
    monkeypatch, tmp_path
):
    """深市 3 只走 cninfo（含 search_org_id），沪市 11 只走 sse；
    日期范围逐基金：有数据的从该基金最新 period 次月 1 日起，
    无数据的从 2023-01-01 起。"""
    _patch_downloads(monkeypatch, tmp_path)
    monthly_df = _monthly_df(
        [
            ["2026-06", "180201", "平安广州广河REIT", 8000, 120000, 1.0, 2.0, "公告"],
            ["2026-03", "508001", "浙商沪杭甬REIT", 7000, 110000, 0.5, 1.5, "公告"],
        ]
    )

    org_calls = []
    monkeypatch.setattr(
        cninfo, "search_org_id", lambda code: org_calls.append(code) or f"org-{code}"
    )
    cninfo_calls = []
    sse_calls = []

    def fake_cninfo_list(code, org_id, date_from, date_to, page_size=100):
        cninfo_calls.append({"code": code, "from": date_from})
        return []

    def fake_sse_list(code, date_from, date_to, page_size=25):
        sse_calls.append({"code": code, "from": date_from})
        return []

    monkeypatch.setattr(cninfo, "list_announcements", fake_cninfo_list)
    monkeypatch.setattr(sse, "list_announcements", fake_sse_list)

    rows = update.fetch_new_monthly(monthly_df)

    assert rows == []
    assert sorted(org_calls) == SZ_CODES
    assert sorted(c["code"] for c in cninfo_calls) == SZ_CODES
    assert sorted(c["code"] for c in sse_calls) == sorted(SH_CODES)

    from_by_code = {c["code"]: c["from"] for c in cninfo_calls + sse_calls}
    assert from_by_code["180201"] == "2026-07-01"
    assert from_by_code["508001"] == "2026-04-01"
    assert all(
        date_from == "2023-01-01"
        for code, date_from in from_by_code.items()
        if code not in ("180201", "508001")
    )


def test_fetch_new_monthly_collects_errors_and_continues(monkeypatch, tmp_path):
    """单基金网络失败 → 错误收集到 errors，其余基金正常返回，不崩溃。"""
    _patch_downloads(monkeypatch, tmp_path)
    monthly_df = _monthly_df(
        [["2026-06", "180201", "平安广州广河REIT", 8000, 120000, 1.0, 2.0, "公告"]]
    )
    monkeypatch.setattr(cninfo, "search_org_id", lambda code: f"org-{code}")
    monkeypatch.setattr(sse, "list_announcements", lambda *a, **k: [])

    def fake_cninfo_list(code, org_id, date_from, date_to, page_size=100):
        if code == "180202":
            raise RuntimeError("网络超时")
        if code == "180201":
            return [_cninfo_item("关于2026年7月主要运营数据的公告", "180201_202607.pdf")]
        return []

    monkeypatch.setattr(cninfo, "list_announcements", fake_cninfo_list)
    monkeypatch.setattr(
        parser_generic, "parse_pdf", lambda path: _monthly_parse_result("2026-07")
    )

    errors = []
    rows = update.fetch_new_monthly(monthly_df, errors=errors)

    assert len(rows) == 1
    assert rows[0]["code"] == "180201"
    assert len(errors) == 1
    assert "180202" in errors[0]


# ---------------------------------------------------------------------------
# fetch_new_quarterly
# ---------------------------------------------------------------------------


def test_fetch_new_quarterly_filters_advisory_and_dedups(monkeypatch, tmp_path):
    """过滤「季度报告」并排除提示性公告；已存在 (code, period) 跳过；
    沪市走 sse，日期范围逐基金（有数据的从最新季度末次日起，无数据从
    2023-01-01 起）。"""
    _patch_downloads(monkeypatch, tmp_path)
    quarterly_df = _quarterly_df(
        [["2026Q2", "508001", "浙商沪杭甬REIT", 17000, None, 4000, 14000, 7000, None, "季报"]]
    )

    sse_calls = []

    def fake_sse_list(code, date_from, date_to, page_size=25):
        sse_calls.append({"code": code, "from": date_from})
        if code != "508001":
            return []
        return [
            _sse_item("浙商沪杭甬REIT 2026年第2季度报告", "q2.pdf"),
            _sse_item("浙商沪杭甬REIT 2026年第3季度报告", "q3.pdf"),
            _sse_item("关于2026年第3季度报告披露的提示性公告", "notice.pdf"),
            _sse_item("浙商沪杭甬REIT 2026年中期报告", "interim.pdf"),
        ]

    monkeypatch.setattr(sse, "list_announcements", fake_sse_list)
    monkeypatch.setattr(cninfo, "search_org_id", lambda code: f"org-{code}")
    monkeypatch.setattr(cninfo, "list_announcements", lambda *a, **k: [])

    downloaded = []
    real_download = sse.download_pdf

    def recording_download(url, dest):
        downloaded.append(dest.name)
        return real_download(url, dest)

    monkeypatch.setattr(sse, "download_pdf", recording_download)

    parsed = {
        "q2.pdf": {
            "period": "2026Q2",
            "revenue_wan": 17000.0,
            "net_profit_wan": 4000.0,
            "cash_distribution_rate": 1.2,
            "distributable_wan": 14000.0,
            "unit_distributable": 0.14,
            "ebitda_wan": 7000.0,
        },
        "q3.pdf": {
            "period": "2026Q3",
            "revenue_wan": 18000.0,
            "net_profit_wan": 4100.0,
            "cash_distribution_rate": 1.3,
            "distributable_wan": 14500.0,
            "unit_distributable": 0.145,
            "ebitda_wan": 7100.0,
        },
    }
    monkeypatch.setattr(
        parser_quarterly, "parse_quarterly_report", lambda path: parsed[path.name]
    )

    rows = update.fetch_new_quarterly(quarterly_df)

    assert len(rows) == 1
    row = rows[0]
    assert row["code"] == "508001"
    assert row["name"] == "浙商沪杭甬REIT"
    assert row["period"] == "2026Q3"
    assert row["revenue_wan"] == 18000.0
    assert row["distributable_wan"] == 14500.0

    # 提示性公告与中期报告既不下载也不解析
    assert sorted(downloaded) == ["q2.pdf", "q3.pdf"]
    from_by_code = {c["code"]: c["from"] for c in sse_calls}
    assert from_by_code["508001"] == "2026-07-01"
    assert all(
        date_from == "2023-01-01"
        for code, date_from in from_by_code.items()
        if code != "508001"
    )


# ---------------------------------------------------------------------------
# fetch_new_completion
# ---------------------------------------------------------------------------


def test_fetch_new_completion_skips_existing_and_appends_new_year(
    monkeypatch, tmp_path
):
    """已存在 (code, year) 跳过、新 year 追加；「摘要」公告被过滤；
    深市走 cninfo；新记录含 year/predicted_wan/actual_wan/completion_pct + code/name。"""
    _patch_downloads(monkeypatch, tmp_path)
    monkeypatch.setattr(update.time, "sleep", lambda s: None)
    existing = [
        {
            "year": 2022,
            "predicted_wan": 62628.76,
            "actual_wan": 47691.19,
            "completion_pct": 76.15,
            "code": "180201",
            "name": "平安广州广河REIT",
        }
    ]
    monkeypatch.setattr(sse, "list_announcements", lambda *a, **k: [])
    monkeypatch.setattr(cninfo, "search_org_id", lambda code: f"org-{code}")

    def fake_cninfo_list(code, org_id, date_from, date_to, page_size=100):
        if code != "180201":
            return []
        return [
            _cninfo_item("平安广州广河REIT 2022年年度报告", "180201_2022.pdf"),
            _cninfo_item("平安广州广河REIT 2025年年度报告", "180201_2025.pdf"),
            _cninfo_item("平安广州广河REIT 2025年年度报告摘要", "180201_2025_summary.pdf"),
        ]

    monkeypatch.setattr(cninfo, "list_announcements", fake_cninfo_list)

    parsed = {
        "180201_2022.pdf": {
            "year": 2022,
            "predicted_wan": 62628.76,
            "actual_wan": 47691.19,
            "completion_pct": 76.15,
        },
        "180201_2025.pdf": {
            "year": 2025,
            "predicted_wan": 70000.0,
            "actual_wan": 68000.0,
            "completion_pct": 97.14,
        },
    }
    parse_calls = []

    def fake_parse(path):
        parse_calls.append(path.name)
        return parsed[path.name]

    monkeypatch.setattr(parser_annual, "parse_annual_completion", fake_parse)

    rows = update.fetch_new_completion(existing)

    assert len(rows) == 1
    row = rows[0]
    assert row["code"] == "180201"
    assert row["name"] == "平安广州广河REIT"
    assert row["year"] == 2025
    assert row["predicted_wan"] == 70000.0
    assert row["actual_wan"] == 68000.0
    assert row["completion_pct"] == 97.14

    # 已存在 year 的公告仍下载解析（判定 year）；「摘要」公告既不下载也不解析
    assert sorted(parse_calls) == ["180201_2022.pdf", "180201_2025.pdf"]


def test_fetch_new_completion_filters_summary_and_notice(monkeypatch, tmp_path):
    """「摘要」「提示性」公告既不下载也不解析；沪市走 sse；
    已存在 PDF 跳过下载但仍解析。"""
    _patch_downloads(monkeypatch, tmp_path)
    monkeypatch.setattr(update.time, "sleep", lambda s: None)
    existing = [
        {
            "year": 2022,
            "predicted_wan": 29081.817072,
            "actual_wan": 25059.791987,
            "completion_pct": 86.17,
            "code": "508018",
            "name": "华夏中国交建REIT",
        }
    ]
    monkeypatch.setattr(cninfo, "search_org_id", lambda code: f"org-{code}")
    monkeypatch.setattr(cninfo, "list_announcements", lambda *a, **k: [])
    # 2025 年报 PDF 已存在 → 跳过下载
    (tmp_path / "508018_2025.pdf").write_bytes(b"%PDF-exists")

    def fake_sse_list(code, date_from, date_to, page_size=25):
        if code != "508018":
            return []
        return [
            _sse_item("华夏中国交建REIT 2025年年度报告", "508018_2025.pdf"),
            _sse_item("华夏中国交建REIT 2025年年度报告摘要", "508018_2025_summary.pdf"),
            _sse_item("关于2025年年度报告披露的提示性公告", "508018_2025_notice.pdf"),
        ]

    monkeypatch.setattr(sse, "list_announcements", fake_sse_list)

    downloaded = []
    real_download = sse.download_pdf

    def recording_download(url, dest):
        downloaded.append(dest.name)
        return real_download(url, dest)

    monkeypatch.setattr(sse, "download_pdf", recording_download)

    parse_calls = []

    def fake_parse(path):
        parse_calls.append(path.name)
        return {
            "year": 2025,
            "predicted_wan": 29081.0,
            "actual_wan": 25059.0,
            "completion_pct": 86.17,
        }

    monkeypatch.setattr(parser_annual, "parse_annual_completion", fake_parse)

    rows = update.fetch_new_completion(existing)

    assert len(rows) == 1
    assert rows[0]["code"] == "508018"
    assert rows[0]["year"] == 2025
    assert downloaded == []
    assert parse_calls == ["508018_2025.pdf"]


def test_fetch_new_completion_collects_errors_and_continues(monkeypatch, tmp_path):
    """单基金列表失败 → 错误收集到 errors，其余基金正常返回，不崩溃。"""
    _patch_downloads(monkeypatch, tmp_path)
    monkeypatch.setattr(update.time, "sleep", lambda s: None)
    existing = [
        {
            "year": 2022,
            "predicted_wan": 62628.76,
            "actual_wan": 47691.19,
            "completion_pct": 76.15,
            "code": "180201",
            "name": "平安广州广河REIT",
        }
    ]
    monkeypatch.setattr(sse, "list_announcements", lambda *a, **k: [])
    monkeypatch.setattr(cninfo, "search_org_id", lambda code: f"org-{code}")

    def fake_cninfo_list(code, org_id, date_from, date_to, page_size=100):
        if code == "180202":
            raise RuntimeError("网络超时")
        if code == "180201":
            return [_cninfo_item("平安广州广河REIT 2025年年度报告", "180201_2025.pdf")]
        return []

    monkeypatch.setattr(cninfo, "list_announcements", fake_cninfo_list)
    monkeypatch.setattr(
        parser_annual,
        "parse_annual_completion",
        lambda path: {
            "year": 2025,
            "predicted_wan": 70000.0,
            "actual_wan": 68000.0,
            "completion_pct": 97.14,
        },
    )

    errors = []
    rows = update.fetch_new_completion(existing, errors=errors)

    assert len(rows) == 1
    assert rows[0]["code"] == "180201"
    assert len(errors) == 1
    assert "180202" in errors[0]


def test_fetch_new_completion_sh_retries_list_failure_three_times(
    monkeypatch, tmp_path
):
    """沪市列表失败 → time.sleep(90) 重试 3 次后成功；沪市基金间 15s 间隔。"""
    _patch_downloads(monkeypatch, tmp_path)
    existing = []
    sleeps = []
    monkeypatch.setattr(update.time, "sleep", lambda s: sleeps.append(s))
    monkeypatch.setattr(cninfo, "search_org_id", lambda code: f"org-{code}")
    monkeypatch.setattr(cninfo, "list_announcements", lambda *a, **k: [])

    calls = {}

    def fake_sse_list(code, date_from, date_to, page_size=25):
        calls[code] = calls.get(code, 0) + 1
        if code == "508018":
            if calls[code] < 3:
                raise RuntimeError("接口限流")
            return [_sse_item("华夏中国交建REIT 2025年年度报告", "508018_2025.pdf")]
        return []

    monkeypatch.setattr(sse, "list_announcements", fake_sse_list)
    monkeypatch.setattr(
        parser_annual,
        "parse_annual_completion",
        lambda path: {
            "year": 2025,
            "predicted_wan": 29081.0,
            "actual_wan": 25059.0,
            "completion_pct": 86.17,
        },
    )

    rows = update.fetch_new_completion(existing)

    assert len(rows) == 1
    assert rows[0]["code"] == "508018"
    assert rows[0]["year"] == 2025
    assert calls["508018"] == 3
    # 3 次尝试间 2 次 90s 等待；沪市 11 只基金间 10 次 15s 间隔
    assert sleeps.count(90) == 2
    assert sleeps.count(15) == 10


# ---------------------------------------------------------------------------
# fetch_market_snapshot
# ---------------------------------------------------------------------------


def _read_fund_shares():
    """读取仓库 data/fund_shares.json 的 shares 映射（14 只）。"""
    path = Path(__file__).resolve().parents[1] / "data" / "fund_shares.json"
    return json.loads(path.read_text(encoding="utf-8"))["shares"]


def _quotes_df(prices):
    """构造 get_realtime_quotes 风格 DataFrame（code→price 映射）。"""
    return pd.DataFrame(
        [
            {
                "code": code,
                "name": f"REIT{code}",
                "price": price,
                "pct_change": 0.0,
                "volume": 0,
                "amount": 0.0,
            }
            for code, price in prices.items()
        ]
    )


def test_fetch_market_snapshot_full_success(tmp_path, monkeypatch):
    """mock get_realtime_quotes 全量 14 只 → 快照 14 条、市值精确
    （价 × 份额 / 10000）、latest 覆盖、写回 SNAPSHOT_PATH。"""
    shares = _read_fund_shares()
    prices = {code: round(7.5 + i * 0.25, 2) for i, code in enumerate(sorted(shares))}
    monkeypatch.setattr(update, "SNAPSHOT_PATH", tmp_path / "market_snapshot.json")
    monkeypatch.setattr(
        update.market_data, "get_realtime_quotes", lambda: _quotes_df(prices)
    )

    result = update.fetch_market_snapshot(shares)

    snapshots = result["snapshots"]
    assert len(snapshots) == 14
    today = date.today().isoformat()
    by_code = {row["code"]: row for row in snapshots}
    for code, price in prices.items():
        row = by_code[code]
        assert row["date"] == today
        assert row["price"] == price
        assert row["market_cap_wan"] == round(price * shares[code] / 10000, 2)
    assert len(result["latest"]) == 14
    assert result["latest"]["180201"]["price"] == prices["180201"]
    # 持久化写回文件
    saved = json.loads((tmp_path / "market_snapshot.json").read_text(encoding="utf-8"))
    assert len(saved["snapshots"]) == 14
    assert saved["latest"]["180201"]["price"] == prices["180201"]


def test_fetch_market_snapshot_partial_failure_keeps_old(tmp_path, monkeypatch):
    """mock 只返回 12 只 → 缺失 2 只沿用旧快照 latest + errors 2 条。"""
    shares = _read_fund_shares()
    old_latest = {
        "508033": {"price": 9.11, "market_cap_wan": 273300.0},
        "508086": {"price": 4.2, "market_cap_wan": 420000.0},
    }
    old_snapshots = [
        {
            "date": "2026-08-03",
            "code": "508033",
            "price": 9.11,
            "market_cap_wan": 273300.0,
        },
        {
            "date": "2026-08-03",
            "code": "508086",
            "price": 4.2,
            "market_cap_wan": 420000.0,
        },
    ]
    (tmp_path / "market_snapshot.json").write_text(
        json.dumps({"snapshots": old_snapshots, "latest": old_latest}),
        encoding="utf-8",
    )
    monkeypatch.setattr(update, "SNAPSHOT_PATH", tmp_path / "market_snapshot.json")
    prices = {
        code: round(7.5 + i * 0.25, 2)
        for i, code in enumerate(sorted(c for c in shares if c not in old_latest))
    }
    monkeypatch.setattr(
        update.market_data, "get_realtime_quotes", lambda: _quotes_df(prices)
    )

    errors = []
    result = update.fetch_market_snapshot(shares, errors=errors)

    # 本次 12 条新行 + 旧快照 2 行；latest 含全部 14 只
    assert len(result["snapshots"]) == 14
    assert len(result["latest"]) == 14
    assert result["latest"]["508033"] == old_latest["508033"]
    assert result["latest"]["508086"] == old_latest["508086"]
    assert len(errors) == 2
    error_codes = {err.split("：")[0] for err in errors}
    assert error_codes == {"508033", "508086"}


def test_fetch_market_snapshot_all_failed_returns_old(tmp_path, monkeypatch):
    """全失败（空行情）→ 返回旧快照内容不变 + errors 14 条，不抛、不写回。"""
    shares = _read_fund_shares()
    old = {
        "snapshots": [
            {
                "date": "2026-08-03",
                "code": "508033",
                "price": 9.11,
                "market_cap_wan": 273300.0,
            }
        ],
        "latest": {"508033": {"price": 9.11, "market_cap_wan": 273300.0}},
    }
    path = tmp_path / "market_snapshot.json"
    path.write_text(json.dumps(old), encoding="utf-8")
    monkeypatch.setattr(update, "SNAPSHOT_PATH", path)
    monkeypatch.setattr(
        update.market_data, "get_realtime_quotes", lambda: _quotes_df({})
    )

    errors = []
    result = update.fetch_market_snapshot(shares, errors=errors)

    assert result == old
    assert len(errors) == 14
    # 文件内容未变
    assert json.loads(path.read_text(encoding="utf-8")) == old


def test_fetch_market_snapshot_first_run_all_failed(tmp_path, monkeypatch):
    """无旧快照 + 全失败 → 返回空结构 {"snapshots": [], "latest": {}} + errors
    14 条，不抛。"""
    shares = _read_fund_shares()
    monkeypatch.setattr(update, "SNAPSHOT_PATH", tmp_path / "market_snapshot.json")
    monkeypatch.setattr(
        update.market_data, "get_realtime_quotes", lambda: _quotes_df({})
    )

    errors = []
    result = update.fetch_market_snapshot(shares, errors=errors)

    assert result == {"snapshots": [], "latest": {}}
    assert len(errors) == 14


def test_update_template_summary_includes_snapshot_updated(tmp_path, monkeypatch):
    """mock fetch_market_snapshot 有变化 → 摘要含 snapshot_updated=True。"""
    path = tmp_path / "tpl.xlsx"
    _build_template(path)
    monkeypatch.setattr(update, "fetch_new_monthly", lambda df, errors=None: [])
    monkeypatch.setattr(update, "fetch_new_quarterly", lambda df, errors=None: [])
    today = date.today().isoformat()
    monkeypatch.setattr(
        update,
        "fetch_market_snapshot",
        lambda shares, errors=None: {
            "snapshots": [
                {
                    "date": today,
                    "code": "180201",
                    "price": 7.56,
                    "market_cap_wan": 529200.0,
                }
            ],
            "latest": {"180201": {"price": 7.56, "market_cap_wan": 529200.0}},
        },
    )

    summary = update.update_template(path)

    assert summary["snapshot_updated"] is True


# ---------------------------------------------------------------------------
# update_template
# ---------------------------------------------------------------------------

STATIC_HEADERS = [
    "基金代码",
    "基金简称",
    "底层资产",
    "区域",
    "里程(km)",
    "上市日期",
    "发行规模(亿元)",
    "特许经营剩余年限\n(截至2026)",
    "资产类型",
]

MONTHLY_HEADERS = [
    "报告期",
    "基金代码",
    "基金简称",
    "通行费收入(万元)",
    "日均自然车流量(辆/日)",
    "通行费收入同比(%)",
    "车流量同比(%)",
    "数据来源/备注",
]

QUARTERLY_HEADERS = [
    "报告期",
    "基金代码",
    "基金简称",
    "营业总收入(万元)",
    "营业成本(万元)",
    "净利润(万元)",
    "可供分配金额(万元)",
    "EBITDA(万元)",
    "基金净资产-NAV(万元)",
    "数据来源/备注",
]


def _build_template(path):
    """构造最小模板：静态 2 行、月度 1 行、季度 1 行。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "静态信息"
    ws.append(STATIC_HEADERS)
    ws.append(["180201", "平安广州广河REIT", "广河高速(广州段)", "华南", 99, "2021-06-07", 91.14, 10, "高速公路"])
    ws.append(["508001", "浙商沪杭甬REIT", "杭徽高速", "华东", 123, "2021-06-07", 43.0, 5, "高速公路"])

    ws = wb.create_sheet("月度数据")
    ws.append(MONTHLY_HEADERS)
    ws.append(["2026-06", "180201", "平安广州广河REIT", 8000, 120000, 1.0, 2.0, "月度运营公告（自动采集）"])

    ws = wb.create_sheet("季度数据")
    ws.append(QUARTERLY_HEADERS)
    ws.append(["2026Q2", "508001", "浙商沪杭甬REIT", 17000, None, 4000, 14000, 7000, None, "季度报告（自动采集）"])

    wb.save(path)


def test_update_template_appends_rows_and_reloads(tmp_path, monkeypatch):
    """主入口：新增行写回对应 Sheet，load 验证行数增加、字段映射正确、
    旧行与表头保留，摘要结构正确。"""
    path = tmp_path / "tpl.xlsx"
    _build_template(path)

    new_monthly = [
        dict(
            _monthly_parse_result("2026-07"),
            code="180201",
            name="平安广州广河REIT",
        )
    ]
    new_quarterly = [
        {
            "code": "508001",
            "name": "浙商沪杭甬REIT",
            "period": "2026Q3",
            "revenue_wan": 18000.0,
            "net_profit_wan": 4100.0,
            "cash_distribution_rate": 1.3,
            "distributable_wan": 14500.0,
            "unit_distributable": 0.145,
            "ebitda_wan": 7100.0,
        }
    ]
    monkeypatch.setattr(update, "fetch_new_monthly", lambda df, errors=None: new_monthly)
    monkeypatch.setattr(
        update, "fetch_new_quarterly", lambda df, errors=None: new_quarterly
    )

    completion_path = tmp_path / "annual_completion.json"
    completion_path.write_text(
        json.dumps({"completion": []}), encoding="utf-8"
    )
    monkeypatch.setattr(
        update, "fetch_new_completion", lambda lst, code_to_name=None, errors=None: []
    )
    monkeypatch.setattr(
        update,
        "fetch_market_snapshot",
        lambda shares, errors=None: {"snapshots": [], "latest": {}},
    )

    summary = update.update_template(path, completion_path=completion_path)

    assert summary == {
        "monthly_added": 1,
        "quarterly_added": 1,
        "completion_added": 0,
        "snapshot_updated": False,
        "errors": [],
    }

    monthly = data_loader.load_monthly(path)
    assert len(monthly) == 2
    row = monthly[(monthly["code"] == "180201") & (monthly["period"] == "2026-07")]
    assert len(row) == 1
    row = row.iloc[0]
    parsed = _monthly_parse_result("2026-07")
    assert row["toll_revenue_wan"] == parsed["toll_revenue_wan"]
    assert row["daily_traffic"] == parsed["daily_traffic"]
    assert row["toll_revenue_yoy"] == parsed["revenue_yoy"]
    assert row["traffic_yoy"] == parsed["traffic_yoy"]
    assert row["source"] == "月度运营公告（自动采集）"
    # 旧行保留
    assert len(monthly[monthly["period"] == "2026-06"]) == 1

    quarterly = data_loader.load_quarterly(path)
    assert len(quarterly) == 2
    qrow = quarterly[
        (quarterly["code"] == "508001") & (quarterly["period"] == "2026Q3")
    ]
    assert len(qrow) == 1
    qrow = qrow.iloc[0]
    assert qrow["total_revenue_wan"] == 18000.0
    assert qrow["net_profit_wan"] == 4100.0
    assert qrow["distributable_wan"] == 14500.0
    assert qrow["ebitda_wan"] == 7100.0
    assert qrow["source"] == "季度报告（自动采集）"


def test_update_template_empty_increment(tmp_path, monkeypatch):
    """空增量（无新公告）→ 摘要全 0，errors 为空，模板数据不变。"""
    path = tmp_path / "tpl.xlsx"
    _build_template(path)

    monkeypatch.setattr(update, "fetch_new_monthly", lambda df, errors=None: [])
    monkeypatch.setattr(update, "fetch_new_quarterly", lambda df, errors=None: [])
    monkeypatch.setattr(
        update,
        "fetch_market_snapshot",
        lambda shares, errors=None: {"snapshots": [], "latest": {}},
    )

    summary = update.update_template(path)

    assert summary == {
        "monthly_added": 0,
        "quarterly_added": 0,
        "completion_added": 0,
        "snapshot_updated": False,
        "errors": [],
    }
    assert len(data_loader.load_monthly(path)) == 1
    assert len(data_loader.load_quarterly(path)) == 1


def test_update_template_writes_back_annual_completion_json(tmp_path, monkeypatch):
    """update_template 经 completion_path 把新完成度记录合并写回
    annual_completion.json（原记录保留 + 新记录追加），摘要 completion_added=1。"""
    path = tmp_path / "tpl.xlsx"
    _build_template(path)

    monkeypatch.setattr(update, "fetch_new_monthly", lambda df, errors=None: [])
    monkeypatch.setattr(update, "fetch_new_quarterly", lambda df, errors=None: [])
    monkeypatch.setattr(
        update,
        "fetch_market_snapshot",
        lambda shares, errors=None: {"snapshots": [], "latest": {}},
    )

    existing = [
        {
            "year": 2022,
            "predicted_wan": 62628.76,
            "actual_wan": 47691.19,
            "completion_pct": 76.15,
            "code": "180201",
            "name": "平安广州广河REIT",
        }
    ]
    new_row = {
        "year": 2025,
        "predicted_wan": 70000.0,
        "actual_wan": 68000.0,
        "completion_pct": 97.14,
        "code": "180201",
        "name": "平安广州广河REIT",
    }
    completion_path = tmp_path / "annual_completion.json"
    completion_path.write_text(
        json.dumps({"completion": existing}, ensure_ascii=False), encoding="utf-8"
    )

    seen = {}

    def fake_fetch(lst, code_to_name=None, errors=None):
        seen["existing"] = lst
        return [new_row]

    monkeypatch.setattr(update, "fetch_new_completion", fake_fetch)

    summary = update.update_template(path, completion_path=completion_path)

    assert summary["completion_added"] == 1
    assert [r["code"] for r in seen["existing"]] == ["180201"]
    written = json.loads(completion_path.read_text(encoding="utf-8"))
    assert written["completion"] == existing + [new_row]
