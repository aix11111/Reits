"""tools.reits_collector.backfill_quarterly_nav 模块（季度 NAV 回填）的单元测试。

季度模板 Sheet 与 market_quarterly.json 的 nav_wan 列/键全空：季报不披露 NAV，
需按「最近年报 NAV」映射回填（每季度行 NAV = 该基金报告年严格早于季度年的
最近一份年报 nav_wan，如 2026Q2→2025 年报值、2025Q1→2024 年报值）。

本测试覆盖：
- nav_wan_for_period 纯函数（mock 数据：2026Q2→2025 值、2025Q1→2024 值、
  无匹配→None；dict 与 list 输入兼容；None 值跳过）
- scan_annual_nav 扫描缓存年报（monkeypatch extract_text）
- update_market_completion 补 None→值、缺失 (code, year) 新增行
- update_template_nav 回填模板季度 Sheet nav_wan 列（openpyxl，不动结构）
- update_market_quarterly_nav 为 market_quarterly.json 每行补 nav_wan 键
"""

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from tools.reits_collector import backfill_quarterly_nav as bq


# ---------------------------------------------------------------------------
# nav_wan_for_period（纯函数）
# ---------------------------------------------------------------------------


def test_nav_wan_for_period_latest_prior_year():
    """2026Q2 → 2025 年报值（2026 年尚无年报，取 <2026 的最近一年）。"""
    annual = {2025: 700438.43, 2024: 734851.28, 2023: 817472.93}

    assert bq.nav_wan_for_period("2026Q2", annual) == 700438.43


def test_nav_wan_for_period_q1_uses_prior_year():
    """2025Q1 → 2024 年报值（2025 年报于次年披露，Q1 时未可用）。"""
    annual = {2025: 700438.43, 2024: 734851.28, 2023: 817472.93}

    assert bq.nav_wan_for_period("2025Q1", annual) == 734851.28


def test_nav_wan_for_period_no_match_returns_none():
    """无早于季度年的年报 → None（2021Q3 前无年报）。"""
    assert bq.nav_wan_for_period("2021Q3", {2022: 860066.41}) is None


def test_nav_wan_for_period_empty_or_none_skipped():
    """空映射 / 值全 None → None；带 None 的记录被跳过。"""
    assert bq.nav_wan_for_period("2026Q2", {}) is None
    assert bq.nav_wan_for_period("2026Q2", {2025: None, 2024: None}) is None


def test_nav_wan_for_period_accepts_record_list():
    """输入兼容 [{"year", "nav_wan"}] 列表（market_completion 行结构）。"""
    records = [
        {"year": 2025, "nav_wan": 700438.43},
        {"year": 2024, "nav_wan": 734851.28},
    ]

    assert bq.nav_wan_for_period("2026Q2", records) == 700438.43
    assert bq.nav_wan_for_period("2025Q1", records) == 734851.28


# ---------------------------------------------------------------------------
# scan_annual_nav
# ---------------------------------------------------------------------------


def test_scan_annual_nav_collects_code_year_nav(monkeypatch, tmp_path):
    """缓存年报 → {(code, report_year): {nav_unit_price, nav_wan}}。

    报告年从标题解析（文件名公告年 ≠ 报告年）；沪市文件名以代码开头直接取 code；
    深市数字文件名按报告标题基金全名唯一匹配 code。无代码/无年份/解析异常跳过。"""
    texts = {
        "180201_annual_2026.pdf": (
            "平安广交投广河高速REIT2025 年年度报告\n"
            "期末不动产基金净资产\n7,004,384,332.39\n"
        ),
        "508001_annual_2026.pdf": (
            "浙商证券沪杭甬杭徽高速封闭式基础设施证券投资基金2025年年度报告\n"
            "期末不动产基金净资产\n2,381,543,029.91\n"
        ),
        "1212581549.PDF": (
            "博时招商蛇口封闭式基础设施证券投资基金2021 年年度报告\n"
            "期末基金净资产\n2,081,195,300.00\n"
        ),
        "508001_annual_2022.pdf": "508001 2022 年年度报告\n期末基金净资产\n3,716,339,700.00\n",
    }

    def fake_extract(path):
        return texts[Path(path).name]

    for name, content in texts.items():
        (tmp_path / name).write_text(content, encoding="utf-8")

    monkeypatch.setattr(bq.parser_annual, "extract_text", fake_extract)

    def fake_code_from_filename(name):
        if name.startswith("180201") or name.startswith("508001"):
            return name[:6]
        return None

    monkeypatch.setattr(bq.market_fetch, "_code_from_filename", fake_code_from_filename)
    monkeypatch.setattr(
        bq.market_fetch, "_report_fund_name", lambda text: "博时招商蛇口"
    )
    monkeypatch.setattr(
        bq.market_fetch, "_match_fund_code", lambda name, funds: "180101"
    )

    funds = [
        {"code": "180101", "name": "博时蛇口产园REIT"},
        {"code": "180201", "name": "平安广州广河REIT"},
        {"code": "508001", "name": "浙商沪杭甬REIT"},
    ]
    nav_map, errors = bq.scan_annual_nav([str(tmp_path)], funds)

    assert nav_map[("180201", 2025)]["nav_wan"] == 700438.43
    assert nav_map[("508001", 2025)]["nav_wan"] == 238154.3
    assert nav_map[("180101", 2021)]["nav_wan"] == 208119.53
    # 508001 2022（文件名公告年 2022，标题报告年 2022）
    assert nav_map[("508001", 2022)]["nav_unit_price"] is None


def test_scan_annual_nav_records_extract_errors(monkeypatch, tmp_path):
    """extract_text 抛异常的 PDF 记录到 errors，不崩溃。"""
    (tmp_path / "corrupt.pdf").write_bytes(b"%PDF-corrupt")

    def fake_extract(path):
        raise RuntimeError("PDF 解析失败")

    monkeypatch.setattr(bq.parser_annual, "extract_text", fake_extract)

    nav_map, errors = bq.scan_annual_nav([str(tmp_path)], [])

    assert nav_map == {}
    assert "corrupt.pdf" in "\n".join(errors)


def test_scan_annual_nav_skips_missing_year(monkeypatch, tmp_path):
    """标题无报告年份的文件跳过，不进 nav_map。"""
    (tmp_path / "508020_annual_2026.pdf").write_text(
        "某基金年度报告\n期末基金净资产\n1,000,000.00\n", encoding="utf-8"
    )

    def fake_extract(path):
        return Path(path).read_text(encoding="utf-8")

    monkeypatch.setattr(bq.parser_annual, "extract_text", fake_extract)

    nav_map, _ = bq.scan_annual_nav([str(tmp_path)], [])

    assert nav_map == {}


# ---------------------------------------------------------------------------
# update_market_completion
# ---------------------------------------------------------------------------


def test_update_market_completion_fills_none_and_adds_rows(tmp_path):
    """已存在行补 None→值；缺失 (code, year) 新增行（完成度字段 None）。"""
    mc_path = tmp_path / "market_completion.json"
    existing = [
        {
            "code": "180201",
            "name": "平安广州广河REIT",
            "year": 2023,
            "predicted_wan": None,
            "actual_wan": None,
            "completion_pct": None,
            "nav_unit_price": 11.6782,
            "nav_wan": None,
        },
        {
            "code": "180201",
            "name": "平安广州广河REIT",
            "year": 2024,
            "predicted_wan": None,
            "actual_wan": None,
            "completion_pct": None,
            "nav_unit_price": 10.4979,
            "nav_wan": None,
        },
    ]
    mc_path.write_text(json.dumps({"completion": existing}), encoding="utf-8")

    nav_map = {
        ("180201", 2023): {"nav_unit_price": 11.6782, "nav_wan": 817472.93},
        ("180201", 2024): {"nav_unit_price": 10.4979, "nav_wan": 734851.28},
        ("180201", 2025): {"nav_unit_price": 10.0063, "nav_wan": 700438.43},
    }
    funds = [
        {"code": "180201", "name": "平安广州广河REIT"},
        {"code": "180202", "name": "华夏越秀高速REIT"},
    ]

    filled, added = bq.update_market_completion(str(mc_path), nav_map, funds)

    assert filled == 2
    assert added == 1

    rows = json.loads(mc_path.read_text(encoding="utf-8"))["completion"]
    by_year = {r["year"]: r for r in rows if r["code"] == "180201"}
    assert by_year[2023]["nav_wan"] == 817472.93
    assert by_year[2024]["nav_wan"] == 734851.28
    assert by_year[2025]["nav_wan"] == 700438.43
    assert by_year[2025]["completion_pct"] is None
    assert len(rows) == 3


def test_update_market_completion_keeps_existing_values(tmp_path):
    """已存在非 None 的 nav_wan 不被覆盖（仅补 None）。"""
    mc_path = tmp_path / "market_completion.json"
    existing = [
        {
            "code": "180201",
            "name": "平安广州广河REIT",
            "year": 2022,
            "predicted_wan": 62628.76,
            "actual_wan": 47691.19,
            "completion_pct": 76.15,
            "nav_unit_price": 12.2867,
            "nav_wan": 860066.41,
        }
    ]
    mc_path.write_text(json.dumps({"completion": existing}), encoding="utf-8")

    nav_map = {("180201", 2022): {"nav_unit_price": 12.2867, "nav_wan": 999999.0}}

    filled, added = bq.update_market_completion(str(mc_path), nav_map, [])

    assert (filled, added) == (0, 0)
    rows = json.loads(mc_path.read_text(encoding="utf-8"))["completion"]
    assert rows[0]["nav_wan"] == 860066.41


def test_update_market_completion_skips_unknown_code(tmp_path):
    """nav_map 中的 code 不在 funds → 不新增行（无法取 name）。"""
    mc_path = tmp_path / "market_completion.json"
    mc_path.write_text(json.dumps({"completion": []}), encoding="utf-8")

    nav_map = {("999999", 2025): {"nav_unit_price": 1.0, "nav_wan": 1.0}}

    filled, added = bq.update_market_completion(str(mc_path), nav_map, [])

    assert (filled, added) == (0, 0)


# ---------------------------------------------------------------------------
# update_template_nav / update_market_quarterly_nav
# ---------------------------------------------------------------------------


def _build_quarterly_template(path):
    """最小模板：季度数据 Sheet 两行（含 NAV 列），表头与真实模板一致。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "季度数据"
    ws.append(
        [
            "报告期",
            "基金代码",
            "基金简称",
            "营业总收入(万元)",
            "营业成本(万元)",
            "净利润(万元)",
            "可供分配金额(万元)",
            "EBITDA(万元)",
            "基金净资产-NAV(万元)",
            "数据来源/备注",
        ]
    )
    ws.append(["2025Q1", "180201", "平安广州广河REIT", 18750, None, 7040, 15110, None, None, "季度报告（自动采集）"])
    ws.append(["2026Q2", "180201", "平安广州广河REIT", 16539, 23044, 3032, 9917, 13618, None, "季度报告（自动采集）"])
    ws.append(["2021Q3", "180201", "平安广州广河REIT", 20607, None, 5506, 15748, None, None, "季度报告（自动采集）"])
    wb.save(path)


def test_update_template_nav_fills_quarterly_column(tmp_path):
    """模板季度 Sheet nav_wan 列按 (code, period) 回填最近年报 NAV；
    无匹配行保持空（None）；表头与其他列不动。"""
    path = tmp_path / "tpl.xlsx"
    _build_quarterly_template(path)

    annual_nav_by_code = {
        "180201": {2025: 700438.43, 2024: 734851.28, 2023: 817472.93}
    }

    filled = bq.update_template_nav(str(path), annual_nav_by_code)

    assert filled == 2

    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb["季度数据"]
    header = [c.value for c in ws[1]]
    assert header[8] == "基金净资产-NAV(万元)"
    values = {ws.cell(row=row, column=9).value for row in range(2, ws.max_row + 1)}
    assert 700438.43 in values
    assert 734851.28 in values
    assert None in values
    # 其余列不受影响
    assert ws.cell(row=2, column=1).value == "2025Q1"
    assert ws.cell(row=2, column=4).value == 18750


def test_update_market_quarterly_nav_adds_nav_wan_key(tmp_path):
    """market_quarterly.json 每行补 nav_wan 键（同映射逻辑），有值行计数正确。"""
    mq_path = tmp_path / "market_quarterly.json"
    quarters = [
        {"code": "180201", "period": "2025Q1", "revenue_wan": 18750.0},
        {"code": "180201", "period": "2026Q2", "revenue_wan": 16539.0},
        {"code": "180201", "period": "2021Q3", "revenue_wan": 20607.0},
        {"code": "508001", "period": "2026Q2", "revenue_wan": 18000.0},
    ]
    mq_path.write_text(json.dumps({"quarters": quarters}), encoding="utf-8")

    annual_nav_by_code = {
        "180201": {2025: 700438.43, 2024: 734851.28, 2023: 817472.93},
        "508001": {2025: 238154.3},
    }

    valued = bq.update_market_quarterly_nav(str(mq_path), annual_nav_by_code)

    assert valued == 3

    rows = json.loads(mq_path.read_text(encoding="utf-8"))["quarters"]
    by_period = {r["period"]: r for r in rows if r["code"] == "180201"}
    assert by_period["2026Q2"]["nav_wan"] == 700438.43
    assert by_period["2025Q1"]["nav_wan"] == 734851.28
    assert by_period["2021Q3"]["nav_wan"] is None
    assert all("nav_wan" in r for r in rows)
    # 508001 也有值
    sh_row = next(r for r in rows if r["code"] == "508001")
    assert sh_row["nav_wan"] == 238154.3
    # 其他键保留
    assert rows[0]["revenue_wan"] == 18750.0
