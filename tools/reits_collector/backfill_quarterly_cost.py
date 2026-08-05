"""季度数据 total_cost_wan 回填脚本（一次性数据迁移）。

高速 REITs KPI 缺数字根因之一是季度模板 Sheet 的 total_cost_wan（营业成本）列
全空：营业成本在季报「4.2.1 不动产项目公司整体财务情况」表的「营业成本/费用」
行，此前解析器未覆盖。本脚本扫描 data/_cache/quarterly_market/ 季度报告 PDF
缓存（沪市文件名以代码开头，深市数字文件名按报告标题基金全名与 market_funds
简称唯一匹配），用 parser_quarterly 解析成本行后：

1. 更新模板 季度数据 Sheet 的「营业成本(万元)」列——openpyxl 按 code+period
   匹配，仅填写解析到值的行（旧格式无成本行保持原样）；
2. 为 data/market_quarterly.json 每行补 total_cost_wan 键（无值为 None）。

用法：python -m tools.reits_collector.backfill_quarterly_cost [template_xlsx]
"""

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook

from tools.reits_collector import market_fetch, parser_quarterly

DEFAULT_TEMPLATE = (
    Path(__file__).resolve().parents[2]
    / "data"
    / "REITsMonitor_数据模板_v1.xlsx"
)
MARKET_QUARTERLY_PATH = (
    Path(__file__).resolve().parents[2] / "data" / "market_quarterly.json"
)
MARKET_FUNDS_PATH = (
    Path(__file__).resolve().parents[2] / "data" / "market_funds.json"
)
CACHE_DIR = (
    Path(__file__).resolve().parents[2] / "data" / "_cache" / "quarterly_market"
)


def _pdf_paths(cache_dir) -> list:
    """缓存目录全部 PDF 文件（大小写后缀兼容），按文件名排序。"""
    return sorted(
        (
            p
            for p in Path(cache_dir).iterdir()
            if p.is_file() and p.suffix.lower() == ".pdf"
        ),
        key=lambda p: p.name,
    )


def scan_cost_map(cache_dir, market_funds, errors=None):
    """扫描季度报告缓存，识别 (code, period) → total_cost_wan（万元）。

    仅收集解析到成本值的报告；旧格式（无 4.2.1「营业成本/费用」行）返回
    None 被跳过，保持模板与 JSON 的既有 None 语义。返回 dict 与错误列表。
    """
    if errors is None:
        errors = []
    szse_funds = [
        f for f in market_funds if str(f.get("code") or "").startswith("180")
    ]
    cost_map = {}
    for path in _pdf_paths(cache_dir):
        try:
            text = parser_quarterly.extract_text(path)
        except Exception as exc:
            errors.append(f"{path.name}：{exc}")
            continue
        period = parser_quarterly._parse_period(text)
        if period is None:
            continue
        code = market_fetch._code_from_filename(path.name)
        if code is None:
            full_name = market_fetch._report_fund_name(text)
            if full_name is None:
                errors.append(f"{path.name}：无法识别基金全名")
                continue
            code = market_fetch._match_fund_code(full_name, szse_funds)
            if code is None:
                errors.append(f"{path.name}：无法唯一匹配基金代码")
                continue
        cost = parser_quarterly._parse_total_cost(text)
        if cost is None:
            continue
        cost_map[(str(code), period)] = cost / 10000.0
    return cost_map, errors


def _load_market_funds(path):
    """读 market_funds.json 的 funds 列表；缺失/损坏返回空列表。"""
    try:
        data = json.loads(Path(path).read_text(encoding="utf-8"))
    except (ValueError, OSError):
        return []
    return data.get("funds") if isinstance(data, dict) else []


def update_template_cost(template_path, cost_map):
    """按 (code, period) 把 total_cost_wan（万元）写入模板季度 Sheet。

    营业成本(万元) 列固定为 E（表头含换行，避免字符串匹配漂移）。仅填写
    解析到值的行；无值行保持原样。返回实际回填行数。
    """
    wb = load_workbook(template_path)
    ws = wb["季度数据"]
    cost_col = None
    for idx, cell in enumerate(ws[1], start=1):
        if "营业成本" in str(cell.value or ""):
            cost_col = idx
            break
    if cost_col is None:
        raise ValueError("模板季度数据 Sheet 未找到「营业成本」列")

    filled = 0
    for row in ws.iter_rows(min_row=2):
        code = row[1].value
        period = row[0].value
        key = (str(code), str(period)) if code is not None and period is not None else None
        if key in cost_map:
            row[cost_col - 1].value = cost_map[key]
            filled += 1
    wb.save(template_path)
    return filled


def update_market_quarterly_cost(mq_path, cost_map):
    """为 market_quarterly.json 每行补 total_cost_wan（万元）键并写回。

    无解析值的行写 None。返回带值（非 None）的行数。
    """
    path = Path(mq_path)
    data = json.loads(path.read_text(encoding="utf-8"))
    quarters = data.get("quarters") or []
    valued = 0
    for row in quarters:
        cost = cost_map.get((str(row.get("code")), str(row.get("period"))))
        row["total_cost_wan"] = cost
        if cost is not None:
            valued += 1
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8"
    )
    return valued


def main(argv=None):
    """CLI：扫描缓存 → 回填模板与 market_quarterly.json，打印摘要。"""
    parser = argparse.ArgumentParser(description="回填季度 total_cost_wan")
    parser.add_argument(
        "template",
        nargs="?",
        default=str(DEFAULT_TEMPLATE),
        help="模板 xlsx 路径（默认 data/REITsMonitor_数据模板_v1.xlsx）",
    )
    args = parser.parse_args(argv)

    market_funds = _load_market_funds(MARKET_FUNDS_PATH)
    cost_map, errors = scan_cost_map(CACHE_DIR, market_funds)
    filled = update_template_cost(args.template, cost_map)
    valued = update_market_quarterly_cost(MARKET_QUARTERLY_PATH, cost_map)

    print(f"扫描识别到成本值 {len(cost_map)} 条")
    print(f"模板季度 Sheet 回填 {filled} 行")
    print(f"market_quarterly.json 有值行 {valued} 行")
    for err in errors:
        print(f"  [错误] {err}")
    return {"scanned": len(cost_map), "filled": filled, "valued": valued}


if __name__ == "__main__":
    main()
