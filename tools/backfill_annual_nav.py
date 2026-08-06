"""按年报年末基金净资产回填季度数据 Sheet 的 NAV 列（一次性数据迁移）。

季报不披露 NAV，而看板季度经营明细表展示的「基金净资产-NAV(万元)」列需用
年报年末值回填。本脚本：

1. 扫描年报缓存（data/_cache/annual/ 及 data/_cache/annual_market/），解析
   每基金每年年末基金净资产 nav_wan（万元），得到年度序列
   {code: {报告年: nav_wan}}（报告年从 PDF 标题「{YYYY} 年年度报告」解析，
   复用 backfill_quarterly_nav.scan_annual_nav）；
2. 加载模板季度数据 Sheet → 纯函数 src.valuation.backfill_nav 重算 NAV 列：
   - 报告期 Q4 行 → 当年年报年末值（报告年 == 季度年）；
   - Q1-Q3 行 → 前向填充（ffill）：最近可得、报告年早于季度年的年报值；
   - 无任何年报值的基金/期间 → 保持空（如 508020 上市 2026-04，无年报）；
3. 以年报年末值为准重算：既有值（可能为期初口径）一律覆盖，openpyxl 仅改
   NAV 列、不动其余列与结构。

用法：python -m tools.backfill_annual_nav [template_xlsx]
"""

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook
import pandas as pd

from src import data_loader
from src.valuation import backfill_nav
from tools.reits_collector import backfill_quarterly_nav as bq

DEFAULT_TEMPLATE = bq.DEFAULT_TEMPLATE
MARKET_QUARTERLY_PATH = bq.MARKET_QUARTERLY_PATH


def build_annual_nav(nav_map: dict) -> dict:
    """{(code, report_year): {nav_unit_price, nav_wan}} → {code: {year: nav_wan}}。

    报告年无 nav_wan（None，早期年报未披露净值）的年份不进年度序列，
    避免 ffill 与 Q4 取到空值。
    """
    annual_nav: dict[str, dict[int, float]] = {}
    for (code, year), fields in nav_map.items():
        nav_wan = fields.get("nav_wan")
        if nav_wan is None:
            continue
        annual_nav.setdefault(str(code), {})[int(year)] = nav_wan
    return annual_nav


def write_nav_back(template_path, backfilled_df) -> int:
    """把重算后的 nav_wan 按 (code, period) 写回模板季度 Sheet 的 NAV 列。

    NAV 列按表头含「NAV」定位（表头含换行，避免字符串匹配漂移）；以年报
    年末值为准重算，故对每个数据行覆盖写入（无值则写 None 清空），说明行
    （不在回填 df 中）保持原样。返回实际写入行数。
    """
    wb = load_workbook(template_path)
    ws = wb["季度数据"]
    nav_col = None
    for idx, cell in enumerate(ws[1], start=1):
        if "NAV" in str(cell.value or ""):
            nav_col = idx
            break
    if nav_col is None:
        raise ValueError("模板季度数据 Sheet 未找到「NAV」列")

    lookup = {
        (str(row.code), str(row.period)): row.nav_wan
        for row in backfilled_df.itertuples()
    }
    written = 0
    for row in ws.iter_rows(min_row=2):
        code = row[1].value
        period = row[0].value
        if code is None or period is None:
            continue
        key = (str(code), str(period))
        if key not in lookup:
            continue
        value = lookup[key]
        row[nav_col - 1].value = None if value is None or pd.isna(value) else value
        written += 1
    wb.save(template_path)
    return written


def update_market_quarterly_nav(mq_path, annual_nav) -> int:
    """按新口径（Q4=当年年报值，Q1-Q3=上年 ffill）重算 market_quarterly.json。

    与 backfill_quarterly_nav.update_market_quarterly_nav（报告年严格早于
    季度年）不同，本函数与模板季度 Sheet 采用同一 backfill_nav 逻辑，避免
    两处口径漂移。返回有值（非 None）行数。
    """
    path = Path(mq_path)
    data = json.loads(path.read_text(encoding="utf-8"))
    quarters = data.get("quarters") or []
    df = pd.DataFrame(quarters)
    if not df.empty:
        df = backfill_nav(df, annual_nav)
    values = df["nav_wan"].tolist() if not df.empty else []
    for row, value in zip(quarters, values):
        row["nav_wan"] = None if value is None or pd.isna(value) else value
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8"
    )
    return sum(1 for v in values if v is not None and not pd.isna(v))


def main(argv=None):
    """CLI：扫描年报缓存 → 构建年度序列 → backfill_nav 重算 → 写回模板与
    market_quarterly.json。"""
    parser = argparse.ArgumentParser(description="按年报年末净资产回填季度 NAV")
    parser.add_argument(
        "template",
        nargs="?",
        default=str(DEFAULT_TEMPLATE),
        help="模板 xlsx 路径（默认 data/REITsMonitor_数据模板_v1.xlsx）",
    )
    args = parser.parse_args(argv)

    market_funds = bq._load_market_funds(bq.MARKET_FUNDS_PATH)
    nav_map, errors = bq.scan_annual_nav(bq.ANNUAL_CACHE_DIRS, market_funds)
    annual_nav = build_annual_nav(nav_map)

    quarterly = data_loader.load_quarterly(args.template)
    before = int(quarterly["nav_wan"].isna().sum())
    backfilled = backfill_nav(quarterly, annual_nav)
    after = int(backfilled["nav_wan"].isna().sum())

    written = write_nav_back(args.template, backfilled)
    mq_valued = update_market_quarterly_nav(MARKET_QUARTERLY_PATH, annual_nav)

    print(f"年报缓存识别净值 {len(nav_map)} 条、基金 {len(annual_nav)} 只")
    print(f"季度 NAV 空值：{before} → {after}")
    print(f"模板季度 Sheet 写入 {written} 行")
    print(f"market_quarterly.json 有值行 {mq_valued} 行")
    if after:
        empty_rows = backfilled[backfilled["nav_wan"].isna()][
            ["code", "period"]
        ].to_string(index=False)
        print("剩余空值行：")
        print(empty_rows)
    for err in errors:
        print(f"  [错误] {err}")
    return {
        "annual_entries": len(nav_map),
        "funds": len(annual_nav),
        "before": before,
        "after": after,
        "written": written,
        "mq_valued": mq_valued,
    }


if __name__ == "__main__":
    main()
